#!/usr/bin/env python3
"""Convertit un fichier Excel (.xlsx/.xlsm) en CSV.

Par défaut:
- 1ère feuille
- encodage UTF-8 avec BOM (plus compatible Excel Windows)
- séparateur virgule
"""
import argparse
import csv
import sys
from pathlib import Path

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Chemin du fichier Excel (.xlsx/.xlsm)")
    ap.add_argument("--sheet", default=None, help="Nom de feuille à convertir (optionnel)")
    ap.add_argument("--out", default=None, help="Chemin du CSV de sortie (optionnel)")
    ap.add_argument(
        "--delimiter",
        default=",",
        choices=[",", ";", "\t"],
        help="Séparateur CSV: ',' ';' ou tabulation (défaut: ',')",
    )
    ap.add_argument(
        "--encoding",
        default="utf-8-sig",
        choices=["utf-8", "utf-8-sig"],
        help="Encodage de sortie (défaut: utf-8-sig pour Excel Windows)",
    )
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("ERREUR: openpyxl n'est pas installé. Installe-le avec: pip install openpyxl")
        sys.exit(1)

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        print(f"ERREUR: fichier introuvable: {xlsx_path}")
        sys.exit(2)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print("ERREUR: feuille non trouvée. Feuilles disponibles:")
            for n in wb.sheetnames:
                print(" -", n)
            sys.exit(3)
        ws = wb[args.sheet]
        sheet_name = args.sheet
    else:
        ws = wb[wb.sheetnames[0]]
        sheet_name = wb.sheetnames[0]

    out_path = Path(args.out).expanduser().resolve() if args.out else xlsx_path.with_suffix(f".{sheet_name}.csv" if args.sheet else ".csv")

    # Write CSV
    with out_path.open("w", newline="", encoding=args.encoding) as f:
        w = csv.writer(f, delimiter=args.delimiter)
        for row in ws.iter_rows(values_only=True):
            # Convert None -> "", keep numbers/strings
            w.writerow([("" if v is None else v) for v in row])

    delim_label = "TAB" if args.delimiter == "\t" else args.delimiter
    print(
        f"OK: {xlsx_path.name}  →  {out_path.name}  "
        f"(feuille: {sheet_name}, séparateur: {delim_label}, encodage: {args.encoding})"
    )

if __name__ == "__main__":
    main()
